"""
LinkedIn Company Insights HTML Extractor
=========================================
Extracts structured data from LinkedIn Insights page HTML dumps and
creates a multi-sheet Excel workbook.

Data extracted:
  - Company overview (name, employee size, industry, HQ, description, tenure)
  - Total employee count & growth trends (6m, 1y, 2y)
  - Functional distribution with employee counts and growth
  - Job openings growth by function
  - Headcount growth by function
  - New hires data
  - AI-generated hiring trends summary
"""

import os
import re
import json
import glob
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Configuration ───────────────────────────────────────────────────────────

HTML_DIR = "/home/jain/Downloads"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "linkedin_company_insights.xlsx")

# ─── Helper functions ────────────────────────────────────────────────────────

def clean_text(text):
    """Collapse whitespace and strip."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()


def parse_growth_cell(cell_text):
    """
    Parse a growth cell like '30%30% increase' or '19%-19% decrease'.
    Returns (percentage_value, direction) e.g. (30, 'increase') or (-19, 'decrease').
    """
    if not cell_text:
        return (None, "")
    cell_text = cell_text.strip()
    # Pattern: number% followed by optional -number% and text
    m = re.search(r'([+-]?\d+)%\s*(increase|decrease|No change)?', cell_text)
    if m:
        val = int(m.group(1))
        direction = m.group(2) or ""
        # If there's a negative sign before the percentage in text like '-19%'
        if '-19%' in cell_text or re.search(r'-\d+%', cell_text):
            # Check if the value itself is negative
            if not cell_text.startswith('-') and '%' in cell_text:
                # The format is '19%-19% decrease' - first is 3m, second is 6m
                pass
        return (val, direction)
    if 'No change' in cell_text:
        return (0, 'No change')
    return (None, "")


def parse_growth_cell_value(cell_text):
    """
    Parse a single growth cell. LinkedIn cells have duplicate content:
    - aria-hidden span: '30%' (visual, always absolute value)
    - visually-hidden span: '30% increase' or '-19% decrease' (screen reader, correct signed value)
    These get concatenated as '30%30% increase' or '19%-19% decrease'.
    
    Returns a clean string like '+30% increase' or '-19% decrease'.
    """
    if not cell_text:
        return ""
    cell_text = cell_text.strip()

    if 'No change' in cell_text or cell_text == '0%':
        return '0% (No change)'

    # Pattern: X%Y% direction — Y carries the actual sign
    m = re.match(r'(-?\d+)%(-?\d+)%\s*(increase|decrease)?', cell_text)
    if m:
        p2 = int(m.group(2))  # The signed value from visually-hidden span
        direction = m.group(3) or ('decrease' if p2 < 0 else 'increase')
        sign = '+' if p2 > 0 else ''
        return f"{sign}{p2}% {direction}"

    # Pattern: single value like '30% increase'
    m = re.match(r'(-?\d+)%\s*(increase|decrease)', cell_text)
    if m:
        pct = int(m.group(1))
        direction = m.group(2)
        sign = '+' if pct > 0 else ''
        return f"{sign}{pct}% {direction}"

    return cell_text


def clean_growth_text(text):
    """
    Clean up merged growth text like '2%2% increase' -> '+2% increase',
    '1%-1% decrease' -> '-1% decrease', '0%No change' -> '0% (No change)'.
    """
    if not text:
        return ""
    text = text.strip()

    if 'No change' in text:
        return '0% (No change)'

    # Pattern: 'X%Y% direction' where Y carries the actual sign
    m = re.match(r'(-?\d+)%(-?\d+)%\s*(increase|decrease)?', text)
    if m:
        p1 = int(m.group(1))
        p2 = int(m.group(2))
        direction = m.group(3) or ('decrease' if p2 < 0 else 'increase')
        # Use p2 as it carries the sign
        pct = p2
        sign = '+' if pct > 0 else ''
        return f"{sign}{pct}% {direction}"

    return text


def extract_company_name_from_title(soup):
    """Extract company name from the <title> tag."""
    title_tag = soup.find('title')
    if title_tag:
        title = title_tag.get_text(strip=True)
        # Remove the LinkedIn suffix
        name = re.sub(r':\s*Insights\s*\|\s*LinkedIn\s*$', '', title).strip()
        return name
    return None


def extract_json_company_data(html, company_name):
    """
    Extract the JSON data block for the main company from the HTML.
    LinkedIn pages contain JSON data for multiple companies (similar cos, etc).
    We need to find the one matching our company name.
    """
    # Find all JSON objects containing company data
    # Pattern: "name":"Company Name" followed by employeeCountRange, etc.
    pattern = r'\{"originalCoverImage"[^}]*"name":"' + re.escape(company_name) + r'"[^}]*"employeeCountRange":\{"start":(\d+),"end":(\d+)[^}]*\}'
    m = re.search(pattern, html)
    if m:
        return {
            'employee_count_start': int(m.group(1)),
            'employee_count_end': int(m.group(2)),
        }

    # Broader search for the company's JSON block
    # Find "name":"<company>" and extract surrounding JSON
    name_escaped = re.escape(company_name)
    pattern2 = r'\{[^}]*"name":"' + name_escaped + r'"[^}]*"employeeCountRange":\{[^}]*"start":(\d+)[^}]*"end":(\d+)[^}]*\}[^}]*\}'
    m2 = re.search(pattern2, html)
    if m2:
        return {
            'employee_count_start': int(m2.group(1)),
            'employee_count_end': int(m2.group(2)),
        }
    return {}


def extract_industry_from_json(html, company_name):
    """Try to find the industry for the main company from JSON data."""
    # Look for industry near the company name in the JSON
    name_escaped = re.escape(company_name)
    # Search for patterns like: "name":"Company"... "industry":[...]
    # This is complex because JSON is nested; use a broader approach

    # Find industries that appear near the company name in the same JSON block
    industries = set()
    # Find all "name":"IndustryName" patterns
    for m in re.finditer(r'"name":"([^"]+)"', html):
        ind_name = m.group(1)
        # Filter out obvious non-industry names
        skip = ['com.linkedin', 'Careers @', 'Life @', 'Learning', 'Find,', 'Unlock',
                'Get qualified', 'Acquire', 'Expand', 'Courses', 'Manage', 'Adamet',
                'PharmSight', 'Axtria', 'Infosys', 'Amazon', 'Accenture',
                'American Express', 'Altimetrik', 'AIS Info', 'Optum',
                'Dual Sphere', 'ChatGPT', 'IBM', 'GSPANN', 'ZS', 'Ernst']
        if any(s in ind_name for s in skip):
            continue
        if len(ind_name) > 3 and len(ind_name) < 100:
            industries.add(ind_name)
    return list(industries)


def extract_headquarters_from_json(html, company_name):
    """Extract headquarters location from JSON data."""
    hq_info = {}
    # Look for address data near the company name
    # Pattern: "headquarter":{..."country":"XX","geographicArea":"...","city":"..."}
    name_escaped = re.escape(company_name)

    # Search for country code near company name
    country_pattern = r'"country":"([A-Z]{2})"'
    geo_pattern = r'"geographicArea":"([^"]+)"'
    city_pattern = r'"city":"([^"]+)"'

    # Find all matches and try to associate with the right company
    # This is heuristic - we look for address data in the vicinity of company name
    return hq_info


def extract_employee_size_text(body_text):
    """
    Extract employee size category like '1K-5K employees', '5,001-10,000 employees',
    or '10,001+ employees'. Tries to find the size associated with the main company
    (not sidebar recommendations) by looking for patterns near 'followers' first,
    then falling back to the first K-format match.
    """
    # Priority 1: Look for employee size right after 'followers' (company header)
    m = re.search(r'followers?\s+([\d,]+-[\d,]+ employees|[0-9]+K-[0-9]+K employees|[\d,]+\+\s*employees)', body_text)
    if m:
        return m.group(1)

    # Priority 2: Look for K-format (more specific to smaller companies): '1K-5K employees', '5K-10K employees'
    m = re.search(r'([0-9]+K-[0-9]+K employees)', body_text)
    if m:
        return m.group(1)

    # Priority 3: Look for '10,001+ employees' that is NOT inside JSON context
    for m in re.finditer(r'([\d,]+\+\s*employees)', body_text):
        # Check if this is a sidebar/JSON reference (preceded by company name + industry)
        before = body_text[max(0, m.start()-100):m.start()]
        if 'account_iq_competitors' not in before and '"text":"' not in before:
            return m.group(1)

    # Priority 4: Any numeric range
    m = re.search(r'([\d,]+-[\d,]+ employees)', body_text)
    if m:
        return m.group(1)

    return ""


def derive_employee_size_from_total(total_employees_str):
    """Derive employee size category from total employee count."""
    if not total_employees_str:
        return ""
    try:
        total = int(total_employees_str.replace(',', ''))
        if total <= 10:
            return "2-10 employees"
        elif total <= 50:
            return "11-50 employees"
        elif total <= 200:
            return "51-200 employees"
        elif total <= 500:
            return "201-500 employees"
        elif total <= 1000:
            return "501-1,000 employees"
        elif total <= 5000:
            return "1,001-5,000 employees"
        elif total <= 10000:
            return "5,001-10,000 employees"
        else:
            return "10,001+ employees"
    except (ValueError, TypeError):
        return ""


def extract_total_employees_and_trends(soup):
    """
    Extract total employee count and growth trends from Table 3.
    Returns dict with total, 6m growth, 1y growth, 2y growth.
    """
    tables = soup.find_all('table')
    result = {'total_employees': None, 'growth_6m': None, 'growth_1y': None, 'growth_2y': None}

    for table in tables:
        rows = table.find_all('tr')
        row_texts = []
        for row in rows:
            cells = row.find_all(['td', 'th'])
            cell_texts = [c.get_text(strip=True) for c in cells]
            row_texts.append(' | '.join(cell_texts))

        full_text = ' '.join(row_texts)
        if 'total employees' in full_text.lower():
            # Table 3 format: first row has numbers, second row has labels
            if len(rows) >= 2:
                val_cells = [c.get_text(strip=True) for c in rows[0].find_all(['td', 'th'])]
                lbl_cells = [c.get_text(strip=True) for c in rows[1].find_all(['td', 'th'])]

                for i, lbl in enumerate(lbl_cells):
                    val = val_cells[i] if i < len(val_cells) else ""
                    if 'total employee' in lbl.lower():
                        result['total_employees'] = val.replace(',', '')
                    elif '6m' in lbl.lower():
                        result['growth_6m'] = clean_growth_text(val)
                    elif '1y' in lbl.lower():
                        result['growth_1y'] = clean_growth_text(val)
                    elif '2y' in lbl.lower():
                        result['growth_2y'] = clean_growth_text(val)
            break

    return result


def extract_functional_distribution(soup):
    """
    Extract functional distribution from tables.
    Returns two lists: current_period (from tables 0/1) and historical (from tables 4/5).
    """
    tables = soup.find_all('table')
    results = {
        'current_period': [],    # 3m/6m growth
        'historical_period': [],  # 6m/1y growth
        'current_header': '',
        'historical_header': '',
    }

    detailed_tables = []  # Tables with 'Function | Number of employees' header

    for table in tables:
        rows = table.find_all('tr')
        if not rows:
            continue
        header_cells = [c.get_text(strip=True) for c in rows[0].find_all(['td', 'th'])]
        header_text = ' | '.join(header_cells)

        if 'Function' in header_text and 'Number of employees' in header_text:
            detailed_tables.append(table)

    for idx, table in enumerate(detailed_tables):
        rows = table.find_all('tr')
        header_cells = [c.get_text(strip=True) for c in rows[0].find_all(['td', 'th'])]
        header_text = ' | '.join(header_cells)

        period_data = []
        for row in rows[1:]:
            cells = [c.get_text(strip=True) for c in row.find_all(['td', 'th'])]
            if len(cells) >= 3:
                row_data = {
                    'function': cells[0],
                    'num_employees': cells[1].replace(',', ''),
                    'pct_headcount': cells[2],
                }
                # Growth columns: each cell represents one time period
                # Column mapping: cells[3]=period1 growth, cells[4]=period2 growth
                if len(cells) >= 5:
                    row_data['growth_period1'] = parse_growth_cell_value(cells[3])
                    row_data['growth_period2'] = parse_growth_cell_value(cells[4])
                elif len(cells) >= 4:
                    row_data['growth_period1'] = parse_growth_cell_value(cells[3])
                    row_data['growth_period2'] = ""

                period_data.append(row_data)

        if idx == 0:  # First detailed table = current period (3m/6m)
            results['current_period'] = period_data
            results['current_header'] = header_text
        elif idx == 1:  # Second = historical (6m/1y)
            results['historical_period'] = period_data
            results['historical_header'] = header_text

    return results


def extract_job_openings_growth(soup):
    """Extract job openings growth by function from Table 2."""
    tables = soup.find_all('table')
    results = []

    for table in tables:
        rows = table.find_all('tr')
        if not rows:
            continue
        header_cells = [c.get_text(strip=True) for c in rows[0].find_all(['td', 'th'])]
        header_text = ' | '.join(header_cells)

        if 'Job openings growth' in header_text:
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(['td', 'th'])]
                if len(cells) >= 3:
                    row_data = {
                        'function': cells[0],
                        'growth_3m': cells[1] if len(cells) > 1 else '',
                        'growth_6m': cells[2] if len(cells) > 2 else '',
                    }
                    results.append(row_data)
            break

    return results


def extract_headcount_growth(soup):
    """Extract headcount growth by function from Table 6."""
    tables = soup.find_all('table')
    results = []

    for table in tables:
        rows = table.find_all('tr')
        if not rows:
            continue
        header_cells = [c.get_text(strip=True) for c in rows[0].find_all(['td', 'th'])]
        header_text = ' | '.join(header_cells)

        if 'Headcount growth' in header_text:
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(['td', 'th'])]
                if len(cells) >= 3:
                    row_data = {
                        'function': cells[0],
                        'growth_6m': cells[1] if len(cells) > 1 else '',
                        'growth_1y': cells[2] if len(cells) > 2 else '',
                    }
                    results.append(row_data)
            break

    return results


def extract_new_hires(soup):
    """Extract new hires information."""
    body = soup.find('body')
    if not body:
        return {}
    body_text = body.get_text(' ', ' ')

    result = {}
    m = re.search(r'New Hires\s+(\w+\s+\d{4})\s+(\d+%)\s+Company-wide\s+(\d+\s*month\s*growth)?', body_text)
    if m:
        result['new_hires_month'] = m.group(1)
        result['new_hires_pct'] = m.group(2)
        result['new_hires_period'] = m.group(3) if m.group(3) else ''

    # Also try to find the numeric value
    m2 = re.search(r'(\d+)\s+New Hires', body_text)
    if m2:
        result['new_hires_count'] = m2.group(1)

    return result


def extract_tenure(body_text):
    """Extract median employee tenure."""
    m = re.search(r'Median employee tenure[^0-9]*([0-9.]+)\s*years?', body_text)
    if m:
        return m.group(1) + ' years'
    return ""


def extract_hiring_trends_insights(body_text):
    """Extract the AI-generated hiring trends summary text."""
    # The insights typically start with patterns like:
    # "Focus on <dept>: The '<dept>' department has seen..."
    # "Growth in <dept>: ..."
    # "Expansion in <dept>: ..."

    patterns = [
        r'(Focus on [^.]+\.[^.]*\.)',
        r'(Growth in [^.]+\.[^.]*\.)',
        r'(Expansion in [^.]+\.[^.]*\.)',
        r'(Increase in [^.]+\.[^.]*\.)',
    ]

    insights = []
    for pattern in patterns:
        for m in re.finditer(pattern, body_text):
            text = m.group(1).strip()
            if text not in insights:
                insights.append(text)

    return insights


def extract_industry_info_from_html(soup):
    """Extract industry from the visible page content."""
    body = soup.find('body')
    if not body:
        return ""
    body_text = body.get_text(' ', ' ')

    # Industry is often shown near the company name on LinkedIn pages
    # We look for known industry patterns
    known_industries = [
        'IT Services and IT Consulting', 'Software Development',
        'Technology, Information and Internet', 'Business Consulting and Services',
        'Pharmaceutical Manufacturing', 'Biotechnology Research',
        'Information Technology & Services', 'Financial Services',
        'Management Consulting', 'Computer Software', 'Outsourcing/Offshoring',
        'Airlines/Aviation', 'Engineering Services', 'Professional Services',
        'Outsourcing and Offshoring Consulting', 'Pharmaceuticals',
        'Biotechnology', 'Airlines and Aviation',
    ]

    for ind in known_industries:
        if ind in body_text:
            return ind
    return ""


def parse_company_page(filepath):
    """
    Main parser for a single LinkedIn Insights HTML page.
    Returns a dictionary with all extracted data.
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        html = f.read()

    soup = BeautifulSoup(html, 'html.parser')
    body = soup.find('body')
    body_text = body.get_text(' ', ' ') if body else ""

    # Basic info
    company_name = extract_company_name_from_title(soup)
    if not company_name:
        company_name = os.path.basename(filepath).replace('_ Insights _ LinkedIn.html', '')

    employee_size = extract_employee_size_text(body_text)
    tenure = extract_tenure(body_text)
    industry = extract_industry_info_from_html(soup)

    # Extract all structured data
    emp_trends = extract_total_employees_and_trends(soup)
    func_dist = extract_functional_distribution(soup)
    job_openings = extract_job_openings_growth(soup)
    headcount_growth = extract_headcount_growth(soup)
    new_hires = extract_new_hires(soup)
    trends_insights = extract_hiring_trends_insights(body_text)

    # Fallback: derive employee size from total employee count if text extraction failed
    if not employee_size and emp_trends.get('total_employees'):
        employee_size = derive_employee_size_from_total(emp_trends['total_employees'])

    # Try to get employee range from JSON
    json_data = extract_json_company_data(html, company_name)

    result = {
        'company_name': company_name,
        'employee_size_category': employee_size,
        'industry': industry,
        'median_tenure': tenure,
        'total_employees': emp_trends.get('total_employees'),
        'growth_6m': emp_trends.get('growth_6m'),
        'growth_1y': emp_trends.get('growth_1y'),
        'growth_2y': emp_trends.get('growth_2y'),
        'employee_count_start': json_data.get('employee_count_start'),
        'employee_count_end': json_data.get('employee_count_end'),
        'new_hires_month': new_hires.get('new_hires_month'),
        'new_hires_count': new_hires.get('new_hires_count'),
        'new_hires_pct': new_hires.get('new_hires_pct'),
        'new_hires_period': new_hires.get('new_hires_period'),
        'trends_insights': trends_insights,
        'functional_distribution_current': func_dist['current_period'],
        'functional_distribution_historical': func_dist['historical_period'],
        'job_openings_growth': job_openings,
        'headcount_growth': headcount_growth,
        'current_func_header': func_dist['current_header'],
        'historical_func_header': func_dist['historical_header'],
    }

    return result


# ─── Excel generation ────────────────────────────────────────────────────────

def write_to_excel(all_data, output_path):
    """Write all extracted data to a multi-sheet Excel workbook."""
    wb = Workbook()

    # ── Styles ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    def style_header_row(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    def style_data_cell(ws, row, col):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = wrap_alignment

    # ── Sheet 1: Company Overview ──
    ws1 = wb.active
    ws1.title = "Company Overview"

    overview_headers = [
        'Company Name', 'Employee Size Category', 'Industry',
        'Total Employees (LinkedIn)', 'Median Employee Tenure',
        'Employee Growth (6m)', 'Employee Growth (1y)', 'Employee Growth (2y)',
        'New Hires Month', 'New Hires Count', 'New Hires %', 'New Hires Period',
        'Employee Count Range (JSON)', 'Hiring Trends Insights'
    ]
    for col, header in enumerate(overview_headers, 1):
        ws1.cell(row=1, column=col, value=header)
    style_header_row(ws1, 1, len(overview_headers))

    for row_idx, data in enumerate(all_data, 2):
        ws1.cell(row=row_idx, column=1, value=data['company_name'])
        ws1.cell(row=row_idx, column=2, value=data['employee_size_category'])
        ws1.cell(row=row_idx, column=3, value=data['industry'])
        ws1.cell(row=row_idx, column=4, value=data['total_employees'])
        ws1.cell(row=row_idx, column=5, value=data['median_tenure'])
        ws1.cell(row=row_idx, column=6, value=data['growth_6m'])
        ws1.cell(row=row_idx, column=7, value=data['growth_1y'])
        ws1.cell(row=row_idx, column=8, value=data['growth_2y'])
        ws1.cell(row=row_idx, column=9, value=data['new_hires_month'])
        ws1.cell(row=row_idx, column=10, value=data['new_hires_count'])
        ws1.cell(row=row_idx, column=11, value=data['new_hires_pct'])
        ws1.cell(row=row_idx, column=12, value=data['new_hires_period'])
        emp_range = f"{data['employee_count_start']}-{data['employee_count_end']}" if data.get('employee_count_start') else ""
        ws1.cell(row=row_idx, column=13, value=emp_range)
        ws1.cell(row=row_idx, column=14, value=' | '.join(data.get('trends_insights', [])))
        for col in range(1, len(overview_headers) + 1):
            style_data_cell(ws1, row_idx, col)

    # Set column widths
    col_widths_1 = [30, 22, 35, 22, 20, 18, 18, 18, 16, 14, 12, 16, 22, 60]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Functional Distribution (Current Period - 3m/6m) ──
    ws2 = wb.create_sheet("Functional Dist (3m-6m)")
    func_headers = ['Company', 'Function', '# Employees', '% of Headcount',
                    '3 Month Growth', '6 Month Growth']
    for col, header in enumerate(func_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header_row(ws2, 1, len(func_headers))

    row_idx = 2
    for data in all_data:
        for func in data.get('functional_distribution_current', []):
            ws2.cell(row=row_idx, column=1, value=data['company_name'])
            ws2.cell(row=row_idx, column=2, value=func.get('function', ''))
            ws2.cell(row=row_idx, column=3, value=func.get('num_employees', ''))
            ws2.cell(row=row_idx, column=4, value=func.get('pct_headcount', ''))
            ws2.cell(row=row_idx, column=5, value=func.get('growth_period1', ''))
            ws2.cell(row=row_idx, column=6, value=func.get('growth_period2', ''))
            for col in range(1, len(func_headers) + 1):
                style_data_cell(ws2, row_idx, col)
            row_idx += 1

    col_widths_2 = [30, 30, 16, 16, 18, 18]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Functional Distribution (Historical - 6m/1y) ──
    ws3 = wb.create_sheet("Functional Dist (6m-1y)")
    for col, header in enumerate(func_headers, 1):
        ws3.cell(row=1, column=col, value=header)
    style_header_row(ws3, 1, len(func_headers))

    # Adjust headers for 6m/1y
    ws3.cell(row=1, column=5, value='6 Month Growth')
    ws3.cell(row=1, column=6, value='1 Year Growth')

    row_idx = 2
    for data in all_data:
        for func in data.get('functional_distribution_historical', []):
            ws3.cell(row=row_idx, column=1, value=data['company_name'])
            ws3.cell(row=row_idx, column=2, value=func.get('function', ''))
            ws3.cell(row=row_idx, column=3, value=func.get('num_employees', ''))
            ws3.cell(row=row_idx, column=4, value=func.get('pct_headcount', ''))
            ws3.cell(row=row_idx, column=5, value=func.get('growth_period1', ''))
            ws3.cell(row=row_idx, column=6, value=func.get('growth_period2', ''))
            for col in range(1, len(func_headers) + 1):
                style_data_cell(ws3, row_idx, col)
            row_idx += 1

    for i, w in enumerate(col_widths_2, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Job Openings Growth ──
    ws4 = wb.create_sheet("Job Openings Growth")
    job_headers = ['Company', 'Function', '3 Month Growth', '6 Month Growth']
    for col, header in enumerate(job_headers, 1):
        ws4.cell(row=1, column=col, value=header)
    style_header_row(ws4, 1, len(job_headers))

    row_idx = 2
    for data in all_data:
        for job in data.get('job_openings_growth', []):
            ws4.cell(row=row_idx, column=1, value=data['company_name'])
            ws4.cell(row=row_idx, column=2, value=job.get('function', ''))
            ws4.cell(row=row_idx, column=3, value=job.get('growth_3m', ''))
            ws4.cell(row=row_idx, column=4, value=job.get('growth_6m', ''))
            for col in range(1, len(job_headers) + 1):
                style_data_cell(ws4, row_idx, col)
            row_idx += 1

    col_widths_4 = [30, 30, 20, 20]
    for i, w in enumerate(col_widths_4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 5: Headcount Growth by Function ──
    ws5 = wb.create_sheet("Headcount Growth")
    hc_headers = ['Company', 'Function', '6 Month Growth', '1 Year Growth']
    for col, header in enumerate(hc_headers, 1):
        ws5.cell(row=1, column=col, value=header)
    style_header_row(ws5, 1, len(hc_headers))

    row_idx = 2
    for data in all_data:
        for hc in data.get('headcount_growth', []):
            ws5.cell(row=row_idx, column=1, value=data['company_name'])
            ws5.cell(row=row_idx, column=2, value=hc.get('function', ''))
            ws5.cell(row=row_idx, column=3, value=hc.get('growth_6m', ''))
            ws5.cell(row=row_idx, column=4, value=hc.get('growth_1y', ''))
            for col in range(1, len(hc_headers) + 1):
                style_data_cell(ws5, row_idx, col)
            row_idx += 1

    col_widths_5 = [30, 30, 20, 20]
    for i, w in enumerate(col_widths_5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        ws.freeze_panes = 'A2'

    # ── Save ──
    wb.save(output_path)
    print(f"✅ Excel file saved to: {output_path}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    # Find all HTML files
    html_files = glob.glob(os.path.join(HTML_DIR, '*Insights*LinkedIn*.html'))
    # Also try without the 'LinkedIn' pattern
    if not html_files:
        html_files = glob.glob(os.path.join(HTML_DIR, '*Insights*.html'))

    if not html_files:
        print(f"❌ No HTML files found in {HTML_DIR}")
        print("   Looking for files matching: *Insights*LinkedIn*.html")
        return

    print(f"📂 Found {len(html_files)} HTML files to process:\n")
    for f in html_files:
        print(f"   - {os.path.basename(f)}")
    print()

    all_data = []
    for filepath in sorted(html_files):
        basename = os.path.basename(filepath)
        print(f"🔄 Processing: {basename} ...")
        try:
            data = parse_company_page(filepath)
            all_data.append(data)
            print(f"   ✅ {data['company_name']}: "
                  f"size={data['employee_size_category']}, "
                  f"total_emp={data['total_employees']}, "
                  f"functions={len(data['functional_distribution_current'])}")
        except Exception as e:
            print(f"   ❌ Error: {e}")
            import traceback
            traceback.print_exc()

    if all_data:
        print(f"\n📊 Writing Excel with {len(all_data)} companies...")
        write_to_excel(all_data, OUTPUT_FILE)
    else:
        print("\n❌ No data extracted.")


if __name__ == '__main__':
    main()
